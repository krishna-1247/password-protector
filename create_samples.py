"""
create_samples.py — one-time script to generate sample files.
Run: python create_samples.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from pypdf import PdfWriter

SAMPLES = [
    ("Rahul.pdf",  "ABCDE1234F"),
    ("Priya.pdf",  "PQRSX5678K"),
    ("Arjun.pdf",  "LMNOP9876Q"),
    ("Sneha.pdf",  "XYZAB5432M"),
    ("Vikram.pdf", "MNPQR7654T"),
]

# ── Create sample passwords.xlsx ──────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Passwords"
ws["A1"] = "PDF File"
ws["B1"] = "Password"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="6C63FF")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

for row in SAMPLES:
    ws.append(row)

ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 18

for d in ("input", "output", "logs"):
    Path(d).mkdir(exist_ok=True)

wb.save("passwords.xlsx")
print(f"passwords.xlsx created with {len(SAMPLES)} entries")

# ── Create sample blank PDFs in input/ ────────────────────────────────────
for name, _ in SAMPLES:
    writer = PdfWriter()
    writer.add_blank_page(width=595, height=842)  # A4
    out = Path("input") / name
    with out.open("wb") as fh:
        writer.write(fh)
    print(f"  Created: {out}")

print("Sample files ready!")
